import os
import time
from typing import Optional, Callable

from django.conf import settings
from openpyxl import Workbook

from smart_meter.models import SmartMeter, PowerMeasurement, GasMeasurement, SolarMeasurement

BATCH_SIZE = 1000

class MeterDataExporter:
    """Service to export meter data to Excel format"""

    def __init__(self, meter: SmartMeter, progress_callback: Optional[Callable[[str], None]] = None):
        """
        Initialize the exporter with a meter and optional progress callback

        :param meter: The SmartMeter instance to export data for
        :param progress_callback: Optional callback function to report progress messages
        """
        self.meter = meter
        self.progress_callback = progress_callback

    def export_to_excel(self, filename: str) -> str:
        """
        Export all measurements for the meter to an Excel file in the media directory

        :param filename: Output xlsx filename (without path)
        :return: Full path to the saved file
        """
        # Construct full file path in media directory
        file_path = os.path.join(settings.MEDIA_ROOT, filename)

        # Ensure directory exists (extract directory from file_path)
        directory = os.path.dirname(file_path)
        if directory:
            os.makedirs(directory, exist_ok=True)

        wb = Workbook(write_only=True)

        self._export_power(wb)
        self._export_gas(wb)
        self._export_solar(wb)

        self._log("Saving workbook...")
        wb.save(file_path)

        return file_path

    def _export_power(self, wb: Workbook) -> None:
        """Export power measurements to a worksheet"""
        ws = wb.create_sheet("Power")

        headers = [
            "timestamp",
            "actual_import",
            "actual_export",
            "total_import_1",
            "total_import_2",
            "total_export_1",
            "total_export_2",
        ]

        ws.append(headers)

        qs = (
            PowerMeasurement.objects
            .filter(meter=self.meter)
            .order_by("id")
        )

        self._write_queryset(
            ws=ws,
            queryset=qs,
            fields=headers,
            title="Power",
        )

    def _export_gas(self, wb: Workbook) -> None:
        """Export gas measurements to a worksheet"""
        ws = wb.create_sheet("Gas")

        headers = [
            "timestamp",
            "actual_gas",
            "total_gas",
        ]

        ws.append(headers)

        qs = (
            GasMeasurement.objects
            .filter(meter=self.meter)
            .order_by("id")
        )

        self._write_queryset(
            ws=ws,
            queryset=qs,
            fields=headers,
            title="Gas",
        )

    def _export_solar(self, wb: Workbook) -> None:
        """Export solar measurements to a worksheet"""
        ws = wb.create_sheet("Solar")

        headers = [
            "timestamp",
            "actual_solar",
            "total_solar",
        ]

        ws.append(headers)

        qs = (
            SolarMeasurement.objects
            .filter(meter=self.meter)
            .order_by("id")
        )

        self._write_queryset(
            ws=ws,
            queryset=qs,
            fields=headers,
            title="Solar",
        )

    def _write_queryset(self, ws, queryset, fields, title):
        """Write queryset data to worksheet in batches"""
        total = queryset.count()

        self._log(f"{title}: exporting {total:,} records")

        if total == 0:
            return

        processed = 0
        last_pk = 0

        while True:
            batch = list(
                queryset
                .filter(pk__gt=last_pk)
                .values_list(*fields, "pk")[:BATCH_SIZE]
            )

            if not batch:
                break

            for row in batch:
                *values, pk = row
                ws.append(values)
                last_pk = pk

            processed += len(batch)

            pct = (processed / total) * 100

            self._log(
                f"{title}: {processed:,}/{total:,} "
                f"({pct:.1f}%)"
            )

            time.sleep(1)

    def _log(self, message: str) -> None:
        """Log a progress message if callback is provided"""
        if self.progress_callback:
            self.progress_callback(message)

